#!/usr/bin/env python
# -*- coding: utf-8 -*-

# AMIU copyleft 2022
# Roberto Marzocchi


'''
Lo script prende in pasto un simil template e fa l'update delle piazzole 
controllando il numero di elementi e nel caso aggiungendo quanto mancante

Segnala tutti i casi in cui i dati non tornano ma c'è già un elemento della stessa tipologia di rifiuto per evitare di fare casino.
E' un caso pericoloso da automatizzare
'''


import os, sys, getopt, re
from dbus import DBusException  # ,shutil,glob
import requests
from requests.exceptions import HTTPError



import json


import inspect, os.path



import openpyxl


import psycopg2
import sqlite3


currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)

sys.path.append(currentdir)

#print(parentdir)
#exit()
#sys.path.append('../')

from credenziali import *
from invio_messaggio import *

#import requests
import datetime

import logging

filename = inspect.getframeinfo(inspect.currentframe()).filename
path = os.path.dirname(os.path.abspath(filename))

giorno_file=datetime.datetime.today().strftime('%Y%m%d%H%M')



#tmpfolder=tempfile.gettempdir() # get the current temporary directory
logfile='{}/log/update_piazzole.log'.format(path)
errorfile='{}/log/error_update_piazzole.log'.format(path)
#if os.path.exists(logfile):
#    os.remove(logfile)

'''logging.basicConfig(
    #handlers=[logging.FileHandler(filename=logfile, encoding='utf-8', mode='w')],
    format='%(asctime)s\t%(levelname)s\t%(message)s',
    #filemode='w', # overwrite or append
    #fileencoding='utf-8',
    #filename=logfile,
    level=logging.DEBUG)
'''


# Create a custom logger
logging.basicConfig(
    level=logging.DEBUG,
    handlers=[
    ]
)

logger = logging.getLogger()

# Create handlers
c_handler = logging.FileHandler(filename=errorfile, encoding='utf-8', mode='w')
f_handler = logging.StreamHandler()
#f_handler = logging.FileHandler(filename=logfile, encoding='utf-8', mode='w')


c_handler.setLevel(logging.WARNING)
f_handler.setLevel(logging.DEBUG)


# Add handlers to the logger
logger.addHandler(c_handler)
logger.addHandler(f_handler)


cc_format = logging.Formatter('%(asctime)s\t%(levelname)s\t%(message)s')

c_handler.setFormatter(cc_format)
f_handler.setFormatter(cc_format)



from invio_messaggio import *



def main():
    
    # Connessione al DB
    logging.info('Connessione al db')
    conn = psycopg2.connect(dbname=db,
                        port=port,
                        user=user,
                        password=pwd,
                        host=host)

    curr = conn.cursor()

    curr1 = conn.cursor()


    curr2 = conn.cursor()

    # Apertura file excel
    logger.info("Apro il file excel")

    xlsx_file='{0}/update_piazzole/template_update.xlsx'.format(path)

    wb = openpyxl.load_workbook(xlsx_file)

    ws = wb["piazzole"]
    logger.debug(ws)


    colonne_leggere=[]
    tipi_elemento=[]
    for x in range(2,3):
        for y in range(2,6):
            if ws.cell(x,y).value:
                logger.debug('{0},{1}'.format(x,y))
                logger.debug(ws.cell(row=x, column=y).value)
                colonne_leggere.append(y)
                tipi_elemento.append(ws.cell(x,y).value)
    


    piazzole=[]
    
    i=1
    for x in ws.values:
        i+=1

    righe=i
    logger.debug('Rows= {}'.format(righe))
    
    for x in range(3,righe):
        for y in range(1,2):
            piazzole.append(ws.cell(x,y).value)


    logger.debug(piazzole)
    logger.debug(colonne_leggere)
    logger.debug(tipi_elemento)
    

    check_commit=0

    # ora devo fare ciclo sulle colonne da leggere
    i=0
    while i< len(colonne_leggere):
        
        # cerco il tipo elemento dalla descrizione scritta sul template
        query_select0='''SELECT tipo_elemento, tipo_rifiuto FROM elem.tipi_elemento te WHERE descrizione ilike %s'''

        try:
            curr.execute(query_select0, (tipi_elemento[i],))
            id_tipo=curr.fetchall()
        except Exception as e:
            logger.error(e)

        id_t=0
        for it in id_tipo:
            id_t=it[0]
            tr=it[1]
        if id_t==0:
            logger.error('Controlla il tipo elemento {} che sembra non esistere'.format(tipi_elemento[i]))
            exit()

        # ciclo sulle piazzole
        p=0
        while p<len(piazzole):
            # leggo il numero di elementi dal file excel
            numero_elementi=int(ws.cell((3+p),colonne_leggere[i]).value)

            # faccio L'INSERT SOLO SE Non CI SONO GIA' ELEMENTI DI QUELLA TIPOLOGIA DI RIFIUTO 
            query_select='''
                select e.id_piazzola, te.descrizione as tipo_elemento, count(e.id_elemento) 
                from elem.elementi e
                join elem.tipi_elemento te on te.tipo_elemento = e.tipo_elemento 
                where e.id_piazzola = %s and te.tipo_rifiuto = %s 
                group by e.id_piazzola, te.descrizione
                '''
            #logger.debug(query_select)
            try:
                curr1.execute(query_select, (piazzole[p],tr,))
                elementi_piazzola=curr1.fetchall()
            except Exception as e:
                logger.error(e)

            if len(elementi_piazzola)==0:
                check_commit=1
                # devo fare insert
                
                insert_query='''
                INSERT INTO elem.elementi
                    (tipo_elemento, 
                    id_piazzola, 
                    id_asta, 
                    x_id_cliente, 
                    privato, peso_reale, peso_stimato,
                    id_utenza,
                    modificato_da,
                    data_ultima_modifica, 
                    percent_riempimento, 
                    freq_stimata, 
                    data_inserimento)
                    values
                    (%s, 
                    %s, 
                    (select id_asta from elem.piazzole where id_piazzola=%s), 
                    '-1'::integer,
                    0, 0, 0,
                    '-1'::integer,
                    '',
                    now(), 
                    90, 
                    3,
                    now()
                    );'''
                
                
                # faccio ciclo per aggiungere il numero di elementi desiderato
                nn=0
                while nn < numero_elementi:
                    logger.info('{4} - Aggiunta bidone {2}/{3} tipo {0} su piazzola {1}'.format(tipi_elemento[i], piazzole[p], (nn+1), numero_elementi, p))              
                    try:
                        curr2.execute(insert_query, (id_t,piazzole[p], piazzole[p]))
                    except Exception as e:
                        logger.error(e)
                    nn+=1
            else:
                # devo dare un WARNING e controllare quelle piazzole a mano perchè pericoloso fare aggiunta da template
                for ee in elementi_piazzola:
                    if tipi_elemento[i]==ee[1] and numero_elementi==ee[2]:
                        logger.info('La piazzola {0} è già ok'.format(ee[0]))
                    else: 
                        w_text= '''La piazzola {0} contiene già {1} elementi di tipo {2}. 
    Sul template ne erano previsti {3} di tipo {4}
    '''.format(ee[0], ee[2], ee[1], numero_elementi, tipi_elemento[i])
                        if input('ATTENZIONE. {} Digita "y" se vuoi procedere ignorando la piazzola\n'.format(w_text)) == "y":
                            logger.warning(w_text)
                        else: 
                            logger.error("Sono uscito alla piazzola {}".format(ee[0]))
                            exit()
            p+=1

        i+=1


    if check_commit==1:
        if input("Sei sicuro di voler continuare ed eseguire il COMMIT (operazione IRREVERSIBILE se non a mano)? [y / other]\n") == "y":
            logger.info("Eseguo il commit")
            conn.commit()
        else: 
            logger.warning("Sono uscito senza fare l COMMIT")

    # check se c_handller contiene almeno una riga 
    error_log_mail(errorfile, 'roberto.marzocchi@amiu.genova.it', os.path.basename(__file__), logger)

    logger.info("chiudo le connessioni")
    curr.close()
    curr1.close()
    curr2.close()
    curr = conn.cursor()

if __name__ == "__main__":
    main()   
