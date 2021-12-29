#!/usr/bin/env python
# -*- coding: utf-8 -*-

# AMIU copyleft 2021
# Roberto Marzocchi

'''
Script per confrontare excel della fase 2 con le piazzole del quartiere
'''


import os,sys, getopt
import inspect, os.path
# da sistemare per Linux
import cx_Oracle

import openpyxl
from pathlib import Path


import xlsxwriter


import psycopg2

import datetime

from urllib.request import urlopen
import urllib.parse

currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)

from credenziali import *
#from credenziali import db, port, user, pwd, host, user_mail, pwd_mail, port_mail, smtp_mail



#libreria per gestione log
import logging


#num_giorno=datetime.datetime.today().weekday()
#giorno=datetime.datetime.today().strftime('%A')

filename = inspect.getframeinfo(inspect.currentframe()).filename
path     = os.path.dirname(os.path.abspath(filename))


giorno_file=datetime.datetime.today().strftime('%Y%m%d')


logfile='{}/log/{}_progettazione_fase2.log'.format(path, giorno_file)

logging.basicConfig(
    handlers=[logging.FileHandler(filename=logfile, encoding='utf-8', mode='a')],
    format='%(asctime)s\t%(levelname)s\t%(message)s',
    #filemode='w', # overwrite or append
    #fileencoding='utf-8',
    #filename=logfile,
    level=logging.INFO)



quartiere=['struppa', 'molassana']


def main():




    # carico i mezzi sul DB PostgreSQL
    logging.info('Connessione al db')
    conn = psycopg2.connect(dbname=db_prog,
                        port=port,
                        user=user,
                        password=pwd,
                        host=host)

    curr = conn.cursor()
    #conn.autocommit = True

    q=0
    while q<len(quartiere):
        xlsx_file = Path('{}/bilaterali/'.format(path), 'fase2.xlsx')
        wbs_obj = openpyxl.load_workbook(xlsx_file) 
        # Read the active sheet:
        #sheet = wbs_obj.active
        sheet = wbs_obj[quartiere[q]]
        logging.debug(sheet)
        print(sheet.max_row)
        piazzole=[]
        codvia=[]
        via=[]
        civico=[]
        riferimento=[]
        suolo_privato=[]
        vocazione_fase2=[]
        ieb=[]
        rsu1=[]
        rsu2=[]
        carta1=[]
        carta2=[]
        multi1=[]
        multi2=[]
        org1=[]
        org2=[]
        note1=[]
        for row in sheet.iter_rows(2, sheet.max_row):
            #print(row[1].value)
            piazzole.append(row[0].value)
            codvia.append(row[1].value)
            via.append(row[2].value)
            civico.append(row[3].value)
            riferimento.append(row[4].value)
            suolo_privato.append(row[5].value)
            vocazione_fase2.append(row[6].value)
            ieb.append(row[7].value)
            rsu1.append(row[8].value)
            rsu2.append(row[9].value)
            carta1.append(row[10].value)
            carta2.append(row[11].value)
            multi1.append(row[12].value)
            multi2.append(row[13].value)
            org1.append(row[14].value)
            org2.append(row[15].value)
            note1.append(row[16].value)

        
        file_output="{0}/bilaterali/output/{1}.xlsx".format(path,quartiere[q])
        logging.info('Scrittura su file {}'.format(file_output))
        workbook = xlsxwriter.Workbook(file_output)
            
        
        w = workbook.add_worksheet('esistenti')
        w.set_tab_color('green')
        
        title1 = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#F9FF33', 'font_size':30, 'valign': 'vcenter', 'center_across': True,'text_wrap': True})

        title2= workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#F9FF33', 'valign': 'vcenter', 'center_across': True,'text_wrap': True})

        text= workbook.add_format({'border': 1, 'valign': 'vcenter', 'text_wrap': True})


        w.set_column(0, 1, 10)
        w.set_column(2, 2, 40)
        w.set_column(3, 3, 10)
        w.set_column(4, 4, 40)
        w.set_column(5, 6, 10) # dalla colonna 5 alla 7
        w.set_column(7, 7, 40)
        w.set_column(8, 8, 10)
        w.set_column(9, 9, 40)
        w.set_column(10, 20, 8)
        w.set_column(21, 21, 40)
        w.set_column(22, 22, 10)

        w.autofilter('A2:W2')

        w.merge_range('A1:J1', 'SIT PROG FASE 1', title1)
        w.merge_range('K1:V1', 'PROGETTAZIONE FASE 2', title1)
        
        w.write(1,0, 'Piazzola', title2)
        w.write(1,1, 'Id_via', title2)
        w.write(1,2, 'Via', title2)
        w.write(1,3, 'Ncivico', title2)
        w.write(1,4, 'Riferimento', title2)
        
        w.write(1,5, 'prog', title2)
        w.write(1,6, 'motivazione', title2)
        w.write(1,7, 'note attuali', title2)
        w.write(1,8, 'suolo privato', title2)
        w.write(1,9, 'elementi', title2)
        w.write(1,10, 'suolo privato', title2)
        w.write(1,11, 'vocazione', title2)
        w.write(1,12, 'IEB', title2)
        w.write(1,13, 'RSU>LU', title2)
        w.write(1,14, 'RSU<LU', title2)
        w.write(1,15, 'CARTA>LU', title2)
        w.write(1,16, 'CARTA<LU', title2)
        w.write(1,17, 'MULTI>LU', title2)
        w.write(1,18, 'MULTI<LU', title2)
        w.write(1,19, 'ORG>LU', title2)
        w.write(1,20, 'ORG<LU', title2)
        w.write(1,21, 'NOTE', title2)
        w.write(1,22, 'DA CAMBIARE', title2)

        query='''select id_piazzola, id_via, nome_via, numero_civico, riferimento, prog, motivazione,note, suolo_privato, 
    string_agg(concat(num, ' x ', tipo_elemento), ',' ) as elementi 
    from
        (select p.id_piazzola, v.id_via, v.nome as nome_via, p.numero_civico, p.riferimento, p.prog, p.motivazione,p.note, p.suolo_privato, 
        count(te.descrizione) as num,  te.descrizione as tipo_elemento 
        from elem.piazzole p 
        left join elem.elementi e 
        on e.id_piazzola = p.id_piazzola 
        left join elem.tipi_elemento te 
        on te.tipo_elemento = e.tipo_elemento 
        join elem.aste a2 
        on a2.id_asta = p.id_asta 
        join topo.vie v 
        on v.id_via = a2.id_via 
        where p.id_asta in 
        (
            select id_asta from elem.aste a where id_quartiere in (
                select id_quartiere from topo.quartieri q where nome ilike '%{0}%'
            )
        ) 
        and (data_eliminazione is null or prog in ('CP', 'CL'))
        group by p.id_piazzola, v.id_via, v.nome, p.numero_civico, p.riferimento, p.prog, p.motivazione, te.descrizione, p.note, p.suolo_privato 
        ) as foo 
    group by id_piazzola, id_via, nome_via, numero_civico, riferimento, prog, motivazione, note, suolo_privato
    order by 2'''.format(quartiere[q])


        check=0
        try:
            curr.execute(query)
            dettagli_piazzola=curr.fetchall()
            check=1
        except Exception as e:
            logging.error(e)
            print('aaa')
        

        k=2
        for u in dettagli_piazzola:
            j=0
            while j<len(u):
                w.write(k,j, u[j], text)
                j+=1 
            if u[0] in piazzole:
                ii=piazzole.index(u[0])
                '''suolo_privato=[]
                vocazione_fase2=[]
                ieb=[]
                rsu1=[]
                rsu2=[]
                carta1=[]
                carta2=[]
                multi1=[]
                multi2=[]
                org1=[]
                org2=[]
                note1=[]'''
                w.write(k,10,suolo_privato[ii], text)
                w.write(k,11,vocazione_fase2[ii], text)
                w.write(k,12,ieb[ii], text)
                w.write(k,13,rsu1[ii], text)
                w.write(k,14,rsu2[ii], text)
                w.write(k,15,carta1[ii], text)
                w.write(k,16,carta2[ii], text)
                w.write(k,17,multi1[ii], text)
                w.write(k,18,multi2[ii], text)
                w.write(k,19,org1[ii], text)
                w.write(k,20,org2[ii], text)
                w.write(k,21,note1[ii], text)
                if vocazione_fase2[ii] == 'E' and u[5] not in ('CL', 'CP'):  
                    w.write(k,22,'x', text)
                #if suolo_privato[ii]=='1' and  vocazione_fase2[ii] not in ('POST', 'E'):
                #    w.write(k,22,'controllare', text)
                if vocazione_fase2[ii] == 'POST' and u[5] is not None: 
                    w.write(k,22,'x', text)
                if ieb[ii] == 'RID' and u[5] not in ('TLR', 'TPR'): 
                    w.write(k,22,'x', text)
                if ieb[ii] == 'STD' and u[5] not in ('TLB', 'TPB'): 
                    w.write(k,22,'x', text)
                if rsu1[ii] in ('2', '3', '4') : 
                    w.write(k,22,'controllare', text)
                if u[8]!=suolo_privato[ii]:
                    w.write(k,22,'controllare suolo', text) 
                if (rsu1[ii]!=org1[ii] or rsu2[ii]!=org2[ii]):
                    w.write(k,22,'controllare', text) 
            #else:
            #    print('no')
            k+=1
            print(k)



        w1 = workbook.add_worksheet('nuove')
        w1.set_tab_color('red')


        w1.set_column(0, 1, 10)
        w1.set_column(2, 2, 40)
        w1.set_column(3, 3, 10)
        w1.set_column(4, 4, 40)
        w1.set_column(5, 15, 10) # dalla colonna 5 alla 7
        w1.set_column(16, 16, 40)
    


        #w1.merge_range('A1:J1', 'SIT PROG FASE 1', title1)
        #w1.merge_range('K1:V1', 'PROGETTAZIONE FASE 2', title1)
        
        w1.write(0,0, 'Piazzola', title2)
        w1.write(0,1, 'Id_via', title2)
        w1.write(0,2, 'Via', title2)
        w1.write(0,3, 'Ncivico', title2)
        w1.write(0,4, 'Riferimento', title2)
        w1.write(0,5, 'suolo_privato', title2)
        w1.write(0,6, 'vocazione', title2)
        w1.write(0,7, 'IEB', title2)
        w1.write(0,8, 'RSU>LU', title2)
        w1.write(0,9, 'RSU<LU', title2)
        w1.write(0,10, 'CARTA>LU', title2)
        w1.write(0,11, 'CARTA<LU', title2)
        w1.write(0,12, 'MULTI>LU', title2)
        w1.write(0,13, 'MULTI<LU', title2)
        w1.write(0,14, 'ORG>LU', title2)
        w1.write(0,15, 'ORG<LU', title2)
        w1.write(0,16, 'NOTE', title2)

        # per le nuove e per le note
        k=0
        i=1
        curr1 = conn.cursor()
        while k <len(piazzole):
            if piazzole[k]=='NUOVA':
                w1.write(i,0,piazzole[k], text)
                w1.write(i,1,codvia[k], text)
                w1.write(i,2,via[k], text)
                w1.write(i,3,civico[k], text)
                w1.write(i,4,riferimento[k], text)
                w1.write(i,5,suolo_privato[k], text)
                w1.write(i,6,vocazione_fase2[k], text)
                w1.write(i,7,ieb[k], text)
                w1.write(i,8,rsu1[k], text)
                w1.write(i,9,rsu2[k], text)
                w1.write(i,10,carta1[k], text)
                w1.write(i,11,carta2[k], text)
                w1.write(i,12,multi1[k], text)
                w1.write(i,13,multi2[k], text)
                w1.write(i,14,org1[k], text)
                w1.write(i,15,org2[k], text)
                w1.write(i,16,note1[k], text)
                i+=1
            else:
                if note1[k] is not None:
                    #print(round(float(piazzole[k]))-20609)
                    if rsu2[k] is not None: 
                        update="update elem.piazzole set note=concat(%s, ' - Lunghezza elementi da verificare') where id_piazzola=%s"
                    else:
                        update="update elem.piazzole set note=%s where id_piazzola=%s"
                    curr1.execute(update, (note1[k],round(float(piazzole[k]))))
                    #print(note1[k])
                elif rsu2[k] is not None: 
                    update="update elem.piazzole set note='Lunghezza elementi da verificare' where id_piazzola=%s"
                    curr1.execute(update, (round(float(piazzole[k])),))
            k+=1


            ########################################################################################
            # da testare sempre prima senza fare i commit per verificare che sia tutto OK
            #conn.commit()
            ########################################################################################


            


        workbook.close()
        q+=1
    curr.close()
    conn.close()






if __name__ == "__main__":
    main()   