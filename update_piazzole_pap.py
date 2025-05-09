#!/usr/bin/env python
# -*- coding: utf-8 -*-

# AMIU copyleft 2022
# Roberto Marzocchi


'''
Lo script prende in pasto un simil template e fa l'update delle piazzole 
controllando il numero di elementi e nel caso aggiungendo quanto mancante

Segnala tutti i casi in cui i dati non tornano ma c'è già un elemento della stessa tipologia di rifiuto per evitare di fare casino.
E' un caso pericoloso da automatizzare
'''


import os, sys, getopt, re
from dbus import DBusException  # ,shutil,glob
import requests
from requests.exceptions import HTTPError



import json


import inspect, os.path



import openpyxl


import psycopg2
import sqlite3


currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)

sys.path.append(currentdir)

#print(parentdir)
#exit()
#sys.path.append('../')

from credenziali import *
from invio_messaggio import *

#import requests
import datetime

import logging

filename = inspect.getframeinfo(inspect.currentframe()).filename
path = os.path.dirname(os.path.abspath(filename))

giorno_file=datetime.datetime.today().strftime('%Y%m%d%H%M')



#tmpfolder=tempfile.gettempdir() # get the current temporary directory
logfile='{}/log/update_piazzole_pap.log'.format(path)
errorfile='{}/log/error_update_piazzole_pap.log'.format(path)
#if os.path.exists(logfile):
#    os.remove(logfile)

'''logging.basicConfig(
    #handlers=[logging.FileHandler(filename=logfile, encoding='utf-8', mode='w')],
    format='%(asctime)s\t%(levelname)s\t%(message)s',
    #filemode='w', # overwrite or append
    #fileencoding='utf-8',
    #filename=logfile,
    level=logging.DEBUG)
'''


# Create a custom logger
logging.basicConfig(
    level=logging.DEBUG,
    handlers=[
    ]
)

logger = logging.getLogger()

# Create handlers
c_handler = logging.FileHandler(filename=errorfile, encoding='utf-8', mode='w')
f_handler = logging.StreamHandler()
#f_handler = logging.FileHandler(filename=logfile, encoding='utf-8', mode='w')


c_handler.setLevel(logging.WARNING)
f_handler.setLevel(logging.DEBUG)


# Add handlers to the logger
logger.addHandler(c_handler)
logger.addHandler(f_handler)


cc_format = logging.Formatter('%(asctime)s\t%(levelname)s\t%(message)s')

c_handler.setFormatter(cc_format)
f_handler.setFormatter(cc_format)



from invio_messaggio import *



def main():
    
    # Connessione al DB
    logging.info('Connessione al db')
    conn = psycopg2.connect(dbname=db,
                        port=port,
                        user=user,
                        password=pwd,
                        host=host)

    curr = conn.cursor()


    # Apertura file excel
    logger.info("Apro il file excel")


    # per ora lo lanciamo a mano con un excel prefissato poi sarebbbe carino renderlo un po' più ingegnerizzato
    #xlsx_file='{0}/update_piazzole/PIAZZOLE_ELEMENTI_PRIVATI.xlsx'.format(path)

    xlsx_file='{0}/update_piazzole/pap_sori.xlsx'.format(path)

    wb = openpyxl.load_workbook(xlsx_file)

    ws = wb["Foglio1"]
    logger.debug(ws)

    piazzole=[]
    macro_categorie=[]
    descrizione_insegna=[]
    nota=[]
    
    i=1
    for x in ws.values:
        i+=1

    righe=i
    logger.debug('Rows= {}'.format(righe))
    
    for x in range(2,righe):
        #for y in range(1,5):
        if ws.cell(x,1).value != None:
            piazzole.append(ws.cell(x,1).value)
            macro_categorie.append(ws.cell(x,2).value)
            descrizione_insegna.append(ws.cell(x,3).value)
            nota.append(ws.cell(x,4).value)


    #logger.debug(piazzole)
    #logger.debug(macro_categorie)
    #logger.debug(descrizione_insegna)
    #logger.debug(nota)    


    # ora devo fare ciclo sulle colonne da leggere
    i=0
    while i< len(piazzole):
        # cerco la macro categoria lo faccio a parte per fare una verifica che tutto sia ok
        select_mc='''select id_macro_categoria from utenze.macro_categorie mc where descrizione = %s'''
        
        try:
            curr.execute(select_mc, (macro_categorie[i],))
            macrocat=curr.fetchall()
        except Exception as e:
            logger.error(e)
            logger.error(select_mc)

        if len(macrocat)==0:
            logger.error('Macro categoria {} non trovata sul DB'.format(macro_categorie[i]))
        else:
            
            for mc in macrocat:
                macro_categoria=mc[0]
        
        curr.close()
        curr = conn.cursor()
        curr1 = conn.cursor()
        
        select_elementi_piazzola='''select id_elemento from elem.elementi where id_piazzola = %s'''
        try:
            curr.execute(select_elementi_piazzola, (piazzole[i],))
            elementi=curr.fetchall()
        except Exception as e:
            logger.error(e)
            logger.error(select_elementi_piazzola)
        
        
        if len(elementi)==0:
            logger.error('Elementi non trovati per piazzola {}'.format(piazzole[i]))
        else:
            for e in elementi: 
                
                insert_elemento_pap=''' INSERT INTO elem.elementi_privati 
                (id_elemento, id_macro_categoria, descrizione, nota)
                values 
                (%s,%s,%s, %s)
                ON CONFLICT (id_elemento)
                DO UPDATE SET
                id_macro_categoria = %s, 
                descrizione= %s, 
                nota=%s               
                '''
                try:
                    curr1.execute(insert_elemento_pap, (e[0], macro_categoria, descrizione_insegna[i], nota[i],
                                                        macro_categoria, descrizione_insegna[i], nota[i] ))
                except Exception as e:
                    logger.error(e)
                    logger.error(insert_elemento_pap)   

                conn.commit()
        i+=1
    
    # check se c_handller contiene almeno una riga 
    error_log_mail(errorfile, 'roberto.marzocchi@amiu.genova.it', os.path.basename(__file__), logger)

    logger.info("chiudo le connessioni")
    curr.close()
    curr1.close()
    conn.close()

if __name__ == "__main__":
    main()   
