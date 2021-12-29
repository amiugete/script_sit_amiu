#!/usr/bin/env python
# -*- coding: utf-8 -*-

# AMIU copyleft 2021
# Roberto Marzocchi

'''
Script per esportare su excel l'elenco delle vie su cui controllare la transitabilità del grafo
'''


import os,sys, getopt
import inspect, os.path
# da sistemare per Linux
import cx_Oracle

import openpyxl
from pathlib import Path


import xlsxwriter


import psycopg2

import datetime

from urllib.request import urlopen
import urllib.parse

currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)

from credenziali import *
#from credenziali import db, port, user, pwd, host, user_mail, pwd_mail, port_mail, smtp_mail



#libreria per gestione log
import logging


#num_giorno=datetime.datetime.today().weekday()
#giorno=datetime.datetime.today().strftime('%A')

filename = inspect.getframeinfo(inspect.currentframe()).filename
path     = os.path.dirname(os.path.abspath(filename))


giorno_file=datetime.datetime.today().strftime('%Y%m%d')


logfile='{}/log/{}_optit_percorsi.log'.format(path, giorno_file)

logging.basicConfig(
    handlers=[logging.FileHandler(filename=logfile, encoding='utf-8', mode='a')],
    format='%(asctime)s\t%(levelname)s\t%(message)s',
    #filemode='w', # overwrite or append
    #fileencoding='utf-8',
    #filename=logfile,
    level=logging.INFO)




def main():




    # carico i mezzi sul DB PostgreSQL
    logging.info('Connessione al db')
    conn = psycopg2.connect(dbname=db,
                        port=port,
                        user=user,
                        password=pwd,
                        host=host)

    curr = conn.cursor()
    conn.autocommit = True

    from pathlib import Path


    piazzole_tot=[]


    mezzi=['Piccolo', 'Porter', 'Porter', 'Isuzu', 'Isuzu', 'Isuzu']
    giri=['Piccolo', 'Porter1', 'Porter2', 'Isuzu1', 'Isuzu2', 'Isuzu3']
    Piccolo=[1,2,3,4,5,6]
    Porter1=[1,3,5,7,9,11]
    Porter2=[2,4,6,8,10,12]
    Isuzu1=[1,4,7,10,13,16]
    Isuzu2=[2,5,8,11,14,17]
    Isuzu3=[3,6,9,12,15,18]
    giorni=['Lunedì', 'Martedì', 'Mercoledì', 'Giovedì', 'Venerdì', 'Sabato']
    color=['NO', 'green', 'red', 'NO', 'green', 'red']
    i=0
    while i < len(mezzi): 
        giro=giri[i]
        j=0
        while j < 6: 
            if giro=='Piccolo':
                num=Piccolo[j]
            elif giro=='Porter1':
                num=Porter1[j]
            elif giro=='Porter2':
                num=Porter2[j]
            elif giro=='Isuzu1':
                num=Isuzu1[j]
            elif giro=='Isuzu2':
                num=Isuzu2[j]
            elif giro=='Isuzu3':
                num=Isuzu3[j]
            xlsx_file = Path('{}/optit/input/sol2/{}/xlsx'.format(path, mezzi[i]), '{}_{}_{}.xlsx'.format(mezzi[i], giorni[j], num))
            wbs_obj = openpyxl.load_workbook(xlsx_file) 
            # Read the active sheet:
            sheet = wbs_obj.active
            logging.debug(sheet)
            #print(sheet.max_row)
            for row in sheet.iter_rows(2, sheet.max_row):
                piazzole_tot.append(row[1].value)
            j+=1
        i+=1

    
    i=0
    while i < len(mezzi): 
        giro=giri[i]
        logging.debug(giro)
        file_giro="{0}/optit/output/sol2/{1}.xlsx".format(path,giro)
        logging.info('Scrittura su file {}'.format(file_giro))
        workbook = xlsxwriter.Workbook(file_giro)
        j=0
        while j < 6: 
            if giro=='Piccolo':
                num=Piccolo[j]
            elif giro=='Porter1':
                num=Porter1[j]
            elif giro=='Porter2':
                num=Porter2[j]
            elif giro=='Isuzu1':
                num=Isuzu1[j]
            elif giro=='Isuzu2':
                num=Isuzu2[j]
            elif giro=='Isuzu3':
                num=Isuzu3[j]
            xlsx_file = Path('{}/optit/input/sol2/{}/xlsx'.format(path, mezzi[i]), '{}_{}_{}.xlsx'.format(mezzi[i], giorni[j], num))
            wbs_obj = openpyxl.load_workbook(xlsx_file) 
            # Read the active sheet:
            sheet = wbs_obj.active
            logging.debug(sheet)
            print(sheet.max_row)
            piazzole=[]
            freq=[]
            for row in sheet.iter_rows(2, sheet.max_row):
                piazzole.append(row[1].value)
                freq.append(row[12].value)




            
            
            w = workbook.add_worksheet(giorni[j])
           
            if color[j]!='NO':
                w.set_tab_color(color[j])
            

            title1 = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#F9FF33', 'font_size':30, 'valign': 'vcenter', 'center_across': True,'text_wrap': True})

            title2= workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#F9FF33', 'valign': 'vcenter', 'center_across': True,'text_wrap': True})

            text= workbook.add_format({'border': 1, 'valign': 'vcenter', 'center_across': True,'text_wrap': True})


            w.set_column(0, 0, 25)
            w.set_column(1, 2, 40)
            w.set_column(3, 3, 40)
            w.set_column(3, 8, 12) # dalla colonna 3 alla 8


            if mezzi[i]=='Porter':
                w.merge_range('A1:F1', '{} - {}'.format( giro, giorni[j]), title1)
            else:
                w.merge_range('A1:H1', '{} - {}'.format( giro, giorni[j]), title1)
            w.write(1,0, 'Piazzola', title2)
            w.write(1,1, 'Indirizzo SIT', title2)
            w.write(1,2, 'Riferimento', title2)
            w.write(1,3, 'Ncivico', title2)
            w.write(1,4, 'rsu240', title2)
            if mezzi[i]!='Porter':
                w.write(1,5, 'rsu660', title2)
                w.write(1,6, 'rsu770', title2)
                w.write(1,7, 'rsu1000', title2)
                w.write(1,8, 'freq_svuot', title2)
                w.write(1,9, 'n_giorni', title2)
            if mezzi[i]=='Porter':
                w.write(1,5, 'sacco pescherie', title2)
                w.write(1,6, 'freq_svuot', title2)
                w.write(1,7, 'n_giorni', title2)

            k=0
            while k <len(piazzole): 
                query='''select nome, riferimento, civico, 
                    rsu240, rsu660, rsu770, rsu1000, saccopescherie 
                    from marzocchir.v_piazzole_percorsi_oregina_castelletto vppoc 
                    where id_piazzola ={}'''.format(piazzole[k])

                check=0
                try:
                    curr.execute(query)
                    dettagli_piazzola=curr.fetchall()
                    check=1
                except Exception as e:
                    logging.error(e)


                #inizializzo gli array
                #ut=[]

                if check==0:
                    w.write(2+k,0,piazzole[k], text)  
                else:
                    w.write(2+k,0,piazzole[k], text)
                    if piazzole[k].strip() == '112':
                        w.write(2+k,1,'Pulizia Via Caffaro', text)
                    if piazzole[k].strip() == '111':
                        w.write(2+k,1,'(30 min circa)', text) 
                    for u in dettagli_piazzola:
                        #logging.debug(vv[0])
                        #ut.append(vv[0])
                        w.write(2+k,1,u[0], text)
                        if u[1] is not None:
                            w.write(2+k,2,u[1], text)
                        w.write(2+k,3,u[2], text)
                        w.write(2+k,4,u[3], text)
                        if mezzi[i]!='Porter':
                            w.write(2+k,5,u[4], text)
                            w.write(2+k,6,u[5], text)
                            w.write(2+k,7,u[6], text)
                            w.write(2+k,8,freq[k], text)
                            w.write(2+k,9,piazzole_tot.count(piazzole_tot[k]), text)
                        if mezzi[i]=='Porter':
                            w.write(2+k,5,u[7], text)
                            w.write(2+k,6,freq[k], text)
                            w.write(2+k,7,piazzole_tot.count(piazzole_tot[k]), text)


                
                #fine piazzola
                k+=1
                
            
            j+=1
            
            #exit()
        i+=1
        workbook.close()






if __name__ == "__main__":
    main()   